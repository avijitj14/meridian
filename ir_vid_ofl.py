#Code for the offline implementation of Smart Mirror.

import os                               #Performing necessary imports
import time
import RPi.GPIO as GPIO
import subprocess
from subprocess import Popen
import logging
#import ibmiotf.device
import Adafruit_GPIO.SPI as SPI
import Adafruit_MCP3008
from threading import Timer
from openpyxl import Workbook
from openpyxl import load_workbook

GPIO.setmode(GPIO.BCM)

SPI_PORT   = 0                        #Defining SPI ports and creating SPI object
SPI_DEVICE = 0
mcp = Adafruit_MCP3008.MCP3008(spi=SPI.SpiDev(SPI_PORT, SPI_DEVICE))

movie=("/home/pi/Videos/Experience The Residence by Etihad Airways.mp4")       #Uploading advertisement video
#black=("/home/pi/Videos/Black Screen for 10 minutes.mp4")
s_no="000000004287dd7f"
loc="GSC GrNoida 1"
schema=['Name','Device','View_No','Location','Hour','Minutes','Day','Date','Month','Year','Duration']
flag_ap=0
hr=0

"""wb = Workbook()                       #To initialize local log file
ws=wb.active
for i in range(1,11):
    c=ws.cell(row=1,column=i)
    c.value=schema[i-1]
ws['N5']='Total Views'
ws['N7']='Last 2 Hours'"""

wb = load_workbook('Log.xlsx')                    #Excel workbook for logging and analytics 
ws = wb.active

def wb_save(hr):                                  #Method for saving
       global views0
       ws['O5']=views
       if(hr==0 and (not flag_ap)):
           hr+=1
       elif(hr==0 and flag_ap):
           ws['O7']=views-views0
       elif(hr==1):
           ws['O7']=views-views0
           hr=0
           views0=views
       elif(hr==2):
           ws['O7']=views-views0
       wb.save('Log.xlsx')
       flag_ap=0

t=Timer(600,wb_save)
t.start()


c=ws['O5']
views=c.value
views0=views
#views=0

present=False


try:
       """options = {
                 "org": "0mvz8a",
                 "type": "raspberry_pi",
                 "id": "rpi_01",
                 "auth-method": "token",
                 "auth-token": "9W@(nZBPSYzAttEinp",
                 "clean-session": False
                 }
       client=ibmiotf.device.Client(options)

       client.connect()"""
       
       while True:
              
             player=False
             present=False
             
             #print mcp.read_adc(0)
             while((mcp.read_adc(0))<50):                            #condition for adc actuation (screen off)
                 #print(mcp.read_adc(0))   
                 #os.system('killall omxplayer.bin')
                 os.system('xset dpms force off')
                 time.sleep(0.1)
                 
             #os.system('killall omxplayer.bin')    
             if((mcp.read_adc(0))>50):                               #condition for adc actuation (screen on) 
                 #print(mcp.read_adc(0))   
                 #os.system('killall omxplayer.bin')
                 omxc=Popen(['omxplayer','-b',movie])
                 time.sleep(15)
                 views+=1
                 player=True
                 view_data=['Etihad Airways',s_no,views,loc,time.strftime("%H",time.localtime()),time.strftime("%M",time.localtime()),time.strftime("%A",time.localtime()),time.strftime("%d",time.localtime()),time.strftime("%m",time.localtime()),time.strftime("%Y",time.localtime()),'15 sec']
                 #myData={'Name' : 'Etihad Airways', 'Device' : s_no, 'View_No':views, 'Location' : loc, 'Hour':time.strftime("%H",time.localtime()),'Minutes':time.strftime("%M",time.localtime()),'Day':time.strftime("%A",time.localtime()), 'Date':time.strftime("%d",time.localtime()),'Month':time.strftime("%m",time.localtime()),'Year':time.strftime("%Y",time.localtime()),'Duration':'15 sec'}
                 #client.publishEvent("log", "json", myData)
                 os.system('killall omxplayer.bin')
                 ws.append(view_data)
                 flag_ap=1

             time.sleep(0.1)
except:                                                              #exception handling
      os.system('killall omxplayer.bin')
      os.system('sleep 1;xset dpms force on')
      player=False
      
      if(flag_ap==1):
             hr=2
             wb_save(hr)
